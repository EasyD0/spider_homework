from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

from article import Article
from write_excel import write_excel
from text_compare import abstract_compare
import numpy as np
import scipy.stats as stats

def get_excel_from_keyword(kw,path):

    a=Article(initialize=False)

    a.get_keywords(kw)
    print(a.Keywords)
    a.get_similar_link()

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures  = [executor.submit(lambda url: Article(url,initialize=True), link) for link in a.Similar_Link]
        similar_article_list = [future.result() for future in as_completed(futures)]

    write_excel(similar_article_list,path=path,sheet_name='相似文献')
    return True
#____________________________________________________________

def get_excel_from_urlname(url='',name='',path='./output.xlsx',requirement='all',timerange='',sim_num=20):
    main_abstract=''
    sim_abstract=[]

    sim_link_list=[]
    ref_link_list=[]
    
    if url:
        art_temp=Article(url=url)
    else:
        art_temp=Article(title=name)

    
    #part1 参考文献
    if requirement=='all' or requirement=='ref':
        art_temp.get_reference_link()
        ref_link_list=art_temp.References_Link
        
        if not ref_link_list:
            print('无参考文献')
        
        else:
            print('有参考文献')
            print(f'将对下面进行爬取{ref_link_list}')
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures  = [executor.submit(lambda url: Article(url,initialize=True), link) for link in ref_link_list]
                article_list = [future.result() for future in as_completed(futures)]
            write_excel(article_list,path=path,sheet_name='参考文献')


    #part2 相似文献
    if requirement == 'all' or requirement == 'sim':
        
        art_temp.get_abstract()
        main_abstract = art_temp.Abstract

        if not main_abstract:
            if input('无法获取主要文章的摘要,是否要退出?') == 'y':
                return

        art_temp.get_similar_link(pages = round(sim_num/25.0+0.5),time_range=timerange,index=True)
        sim_link_list = art_temp.Similar_Link[0:min(sim_num,len(art_temp.Similar_Link))]

        if not sim_link_list:
            print('无相似文献')
        
        else:
            print('有相似文献')
            print(f'将对如下关键词搜素{art_temp.Keywords}')
            print(f'将对下面进行爬取{sim_link_list}')
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures  = [executor.submit(lambda url: Article(url,initialize=True), link) for link in sim_link_list]
                sim_article_list = [future.result() for future in as_completed(futures)]
            write_excel(sim_article_list,path=path,sheet_name='相似文献')

    #TODO: 文本对比
            sim_abstract=[paper.Abstract for paper in sim_article_list]


            scores=abstract_compare(main_abstract, sim_abstract)#获取相似分数
            score_ranking=(np.argsort(-np.array(scores))+1).tolist()

            #写入excel
            workbook = openpyxl.load_workbook(path)
            sheet = workbook['相似文献']
            
            sheet.cell(row=1, column=9, value='余弦相似度')
            [sheet.cell(row=2+i, column=9, value=scores[i]) for i in range(len(scores))]
            sheet.cell(row=1, column=10, value='相似度排名')
            [sheet.cell(row=2+i, column=10, value=score_ranking[i]) for i in range(len(score_ranking))]

            workbook.save(path)

            return sim_article_list, scores, score_ranking
    
    return [],[]


