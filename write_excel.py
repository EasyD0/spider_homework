from article import Article
# import xlsxwriter
import openpyxl
import os


def write_excel(article_list, path, sheet_name):
    title_list = [article.Title for article in article_list]
    link_list = [article.Url for article in article_list]
    abstract_list = [article.Abstract for article in article_list]
    abstract_translation_list = [article.Abstract_Translation for article in article_list]
    date_list = [article.Date for article in article_list]
    journal_name_list = [article.Journal_Name for article in article_list]
    impact_list = [article.Impact_Factor for article in article_list]
    doi_list = [article.Doi for article in article_list]

    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        workbook.save(path)

    workbook = openpyxl.load_workbook(path)
    try:
        workbook.remove(workbook['Sheet'])
    except:
        pass

    sheet = workbook.create_sheet(sheet_name)

    [sheet.cell(row=i + 2, column=1, value=title_list[i]) for i in range(len(title_list))]
    [sheet.cell(row=i + 2, column=2, value=link_list[i]) for i in range(len(link_list))]
    [sheet.cell(row=i + 2, column=3, value=abstract_list[i]) for i in range(len(abstract_list))]
    [sheet.cell(row=i + 2, column=4, value=abstract_translation_list[i]) for i in range(len(abstract_translation_list))]
    [sheet.cell(row=i + 2, column=5, value=date_list[i]) for i in range(len(date_list))]
    [sheet.cell(row=i + 2, column=6, value=journal_name_list[i]) for i in range(len(journal_name_list))]
    [sheet.cell(row=i + 2, column=7, value=impact_list[i]) for i in range(len(impact_list))]
    [sheet.cell(row=i + 2, column=8, value=doi_list[i]) for i in range(len(doi_list))]

    row_list = ['标题', '链接', '摘要', '摘要翻译', '日期', '期刊名/会议名', '影响因子', 'doi']
    [sheet.cell(row=1, column=i + 1, value=row_list[i]) for i in range(len(row_list))]
    workbook.save(path)

    '''
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.write_column(1, 0, title_list)
    worksheet.write_column(1, 1, link_list)
    worksheet.write_column(1, 2, abstract_list)
    worksheet.write_column(1, 3, abstract_translation_list)
    worksheet.write_column(1, 4, date_list)
    worksheet.write_column(1, 5, journal_name_list)
    worksheet.write_column(1, 6, impact_list)
    worksheet.write_column(1, 7, doi_list)
    worksheet.write_row(0,0,['标题','链接','摘要','摘要翻译','日期','期刊名/会议名','影响因子','doi'])
    workbook.close()
    '''
