import openpyxl
import os


def write_excel(path, sheet_name):
    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        workbook.save(path)

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet(sheet_name)
    try:
        workbook.remove(workbook['Sheet'])
    except:
        pass

    sheet.cell(row=1, column=1, value='测试')
    workbook.save(path)


if __name__ == '__main__':
    path = './out.xlsx'
    write_excel(path, '相似文献')
    write_excel(path, '参考文献')
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['相似文献']
    sheet.cell(row=1, column=2, value='测试')
    workbook.save(path)
